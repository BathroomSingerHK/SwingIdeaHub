import math
import warnings
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional

import numpy as np
import pandas as pd
import yfinance as yf
from matplotlib.colors import LinearSegmentedColormap, to_hex

# IMPORTANT: You must ensure this file exists in the same folder
try:
    from update_volatility_360d import update_volatility_360d_column
except ImportError:
    # Fallback if the module isn't found (for testing in isolation)
    def update_volatility_360d_column(path):
        pass

BASE_DIR = Path(__file__).resolve().parent
VOL_COLUMN = "VOLATILITY_360D"

warnings.simplefilter("ignore", category=FutureWarning)


def detect_ticker_column(df: pd.DataFrame) -> str:
    if df.empty:
        raise ValueError("sector_etf_stock_list.xlsx is empty.")

    candidates = [c for c in df.columns if isinstance(c, str)]
    if not candidates:
        raise ValueError("No header found.")

    key_words = ("ticker", "symbol", "code")
    header_match_cols = []
    for col in candidates:
        lower = col.strip().lower()
        if any(k in lower for k in key_words):
            header_match_cols.append(col)

    def is_ticker_like(s: str) -> bool:
        if not isinstance(s, str): return False
        s = s.strip()
        if not s: return False
        if any(ord(ch) > 127 for ch in s): return False
        if len(s) > 15: return False
        # Relaxed check: allow alphanum and standard ticker symbols
        return all(ch.isalnum() or ch in ".-=^+" for ch in s)

    def column_score(col: str) -> float:
        series = df[col].astype(str).fillna("")
        if series.empty: return 0.0
        sample = series.head(200)
        vals = [v for v in sample if v and v.lower() != "nan"]
        if not vals: return 0.0
        like = sum(1 for v in vals if is_ticker_like(v))
        return like / max(1, len(vals))

    scored = [(col, column_score(col)) for col in candidates if pd.api.types.is_object_dtype(df[col])]
    scored.sort(key=lambda x: (x[0] in header_match_cols, x[1]), reverse=True)
    if scored and scored[0][1] > 0:
        return scored[0][0]

    return str(candidates[0])


def read_tickers_from_excel(path: str = "sector_etf_stock_list.xlsx", df: Optional[pd.DataFrame] = None) -> List[str]:
    if df is None:
        df = pd.read_excel(path)
    col = detect_ticker_column(df)
    raw = df[col].astype(str).str.strip().replace({"": np.nan}).dropna().unique().tolist()

    def is_ticker_like(s: str) -> bool:
        if not isinstance(s, str): return False
        s = s.strip()
        if not s: return False
        if any(ord(ch) > 127 for ch in s): return False
        if len(s) > 20: return False  # Increased length allowance slightly
        return all(ch.isalnum() or ch in ".-=^+" for ch in s)

    tickers = [t for t in raw if is_ticker_like(t)]
    if not tickers:
        raise ValueError("No valid tickers found in Excel.")
    return tickers


def read_column_b_mapping(path: str = "sector_etf_stock_list.xlsx", df: Optional[pd.DataFrame] = None) -> dict:
    if df is None:
        df = pd.read_excel(path)
    ticker_col = detect_ticker_column(df)
    if df.shape[1] < 2: return {}

    b_col_idx = 1
    mapping = {}
    for idx, row in df.iterrows():
        ticker_raw = str(row.get(ticker_col, "")).strip()
        if not ticker_raw or ticker_raw.lower() == "nan": continue

        # We don't enforce strict ticker checks on the key here to ensure we capture all mappings
        # but we generally expect the key to match what we read in read_tickers_from_excel

        b_value = row.iloc[b_col_idx]
        if pd.notna(b_value):
            mapping[ticker_raw] = str(b_value).strip()
        else:
            mapping[ticker_raw] = ticker_raw
    return mapping


def read_column_a_mapping(path: str = "sector_etf_stock_list.xlsx", df: Optional[pd.DataFrame] = None) -> dict:
    if df is None:
        df = pd.read_excel(path)
    ticker_col = detect_ticker_column(df)
    if df.shape[1] < 1: return {}

    a_col_idx = 0
    mapping = {}
    for idx, row in df.iterrows():
        ticker_raw = str(row.get(ticker_col, "")).strip()
        if not ticker_raw or ticker_raw.lower() == "nan": continue

        a_value = row.iloc[a_col_idx]
        if pd.notna(a_value):
            mapping[ticker_raw] = str(a_value).strip()
        else:
            mapping[ticker_raw] = ""
    return mapping


def download_prices(tickers: List[str], period: str = "3mo") -> pd.DataFrame:
    # 1. Attempt Bulk Download
    print(f"[INFO] Downloading data for {len(tickers)} tickers...")
    try:
        data = yf.download(tickers, period=period, auto_adjust=True, progress=False, group_by="column")
    except Exception as e:
        print(f"[WARN] Bulk download failed: {e}")
        data = pd.DataFrame()

    prices = pd.DataFrame()

    if not data.empty:
        if isinstance(data.columns, pd.MultiIndex):
            if "Close" in data.columns.get_level_values(0):
                # yfinance structure varies, sometimes level 0 is Close, sometimes level 1
                # If multiindex is (Price, Ticker) or (Ticker, Price)
                # safe way:
                prices = data.xs('Close', axis=1, level=1, drop_level=True) if 'Close' not in data.columns else data[
                    'Close']
            else:
                # Fallback
                prices = data["Close"].copy() if "Close" in data else pd.DataFrame()
        else:
            # Single ticker case
            colname = "Close" if "Close" in data.columns else list(data.columns)[0]
            one = data[[colname]].copy()
            if len(tickers) == 1:
                one.columns = [tickers[0]]
            prices = one

    # 2. Fallback for missing
    missing_cols = [t for t in tickers if t not in prices.columns]
    if missing_cols:
        print(f"[INFO] Retrying {len(missing_cols)} missing tickers individually...")
        series_list = []
        for tk in missing_cols:
            try:
                tdf = yf.download(tk, period=period, auto_adjust=True, progress=False)
                if not tdf.empty and "Close" in tdf.columns and not tdf["Close"].dropna().empty:
                    s = tdf["Close"].copy()
                    s.name = tk
                    series_list.append(s)
            except Exception:
                pass

        if series_list:
            prices_fallback = pd.concat(series_list, axis=1).sort_index()
            prices = pd.concat([prices, prices_fallback], axis=1)

    # 3. [FIX] Force all requested tickers to be in the columns
    # This ensures that even if YFinance didn't return data, the column exists (as NaNs)
    # so the ticker appears in the final heatmap.
    prices = prices.reindex(columns=tickers)

    return prices


def compute_daily_returns(prices: pd.DataFrame) -> pd.DataFrame:
    return prices.pct_change() * 100.0


def latest_20_days_transposed(returns_pct: pd.DataFrame) -> pd.DataFrame:
    # Drop rows (Dates) where ALL tickers are NaN, but keep the Columns (Tickers)
    last = returns_pct.tail(40).dropna(axis=0, how="all").tail(20).copy()
    last.index = last.index.strftime("%Y-%m-%d")
    df = last.T
    # Note: We sort columns (Dates) reverse chronological
    cols_sorted = sorted(df.columns, reverse=True)
    df = df[cols_sorted]
    return df


def prepare_vol_series(meta_df: pd.DataFrame, ticker_col: str, index: pd.Index) -> pd.Series:
    meta_copy = meta_df.copy()
    meta_copy[ticker_col] = meta_copy[ticker_col].astype(str).str.strip()
    if VOL_COLUMN not in meta_copy.columns:
        return pd.Series(np.nan, index=index, dtype=float)
    vol_series = (
        meta_copy.set_index(meta_copy[ticker_col])[VOL_COLUMN]
        .pipe(pd.to_numeric, errors="coerce")
        .reindex(index)
    )
    return vol_series


def prepare_mu_series(meta_df: pd.DataFrame, ticker_col: str, index: pd.Index) -> pd.Series:
    meta_copy = meta_df.copy()
    meta_copy[ticker_col] = meta_copy[ticker_col].astype(str).str.strip()
    if "MU_1D" not in meta_copy.columns:
        return pd.Series(np.nan, index=index, dtype=float)
    mu_series = (
        meta_copy.set_index(meta_copy[ticker_col])["MU_1D"]
        .pipe(pd.to_numeric, errors="coerce")
        .reindex(index)
    )
    return mu_series


def compute_zs_live(df_pct: pd.DataFrame, vol_series: pd.Series) -> pd.Series:
    if df_pct.empty: return pd.Series(dtype=float)
    pct_today = df_pct.iloc[:, 0].astype(float)
    aligned_vol = vol_series.reindex(df_pct.index)
    zs_live = pd.Series(np.nan, index=df_pct.index, dtype=float)
    mask_valid = (aligned_vol > 0) & pct_today.notna()
    zs_live[mask_valid] = pct_today[mask_valid] * 16.0 / aligned_vol[mask_valid]
    return zs_live.replace([np.inf, -np.inf], np.nan)


def compute_zs_5d(df_pct: pd.DataFrame, vol_series: pd.Series, mu_series: pd.Series) -> pd.Series:
    if df_pct.empty or df_pct.shape[1] < 5:
        return pd.Series(np.nan, index=df_pct.index, dtype=float)
    last5_pct = df_pct.iloc[:, :5].astype(float)
    last5_dec = last5_pct / 100.0
    R_5d = (1.0 + last5_dec).prod(axis=1) - 1.0
    sigma_1d = vol_series.reindex(df_pct.index) / (16.0 * 100.0)
    mu_daily_dec = mu_series.reindex(df_pct.index) / 100.0
    zs_5d = pd.Series(np.nan, index=df_pct.index)
    mask_valid_5d = (sigma_1d.notna() & (sigma_1d > 0) & R_5d.notna() & mu_daily_dec.notna())
    zs_5d[mask_valid_5d] = (R_5d[mask_valid_5d] - 5.0 * mu_daily_dec[mask_valid_5d]) / (
            sigma_1d[mask_valid_5d] * math.sqrt(5.0))
    return zs_5d.replace([np.inf, -np.inf], np.nan)


def build_colormap() -> LinearSegmentedColormap:
    def pos(v: float) -> float: return (v + 10.0) / 20.0

    anchor_points = [(pos(-10.0), "#FF0000"), (pos(-1.0), "#C62828"), (pos(0.0), "#F5F5F5"), (pos(1.0), "#0FA84C"),
                     (pos(10.0), "#00FF00")]
    return LinearSegmentedColormap.from_list("sector_returns_5anchors", anchor_points, N=256)


def to_html_heatmap(df_pct: pd.DataFrame, html_path: str, ticker_to_colb: dict = None, ticker_to_cola: dict = None,
                    zs_live: Optional[pd.Series] = None, zs_5d: Optional[pd.Series] = None) -> None:
    cmap = build_colormap()
    if ticker_to_colb is None: ticker_to_colb = {}
    if ticker_to_cola is None: ticker_to_cola = {}

    html = []
    timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    html.append("<html><head><meta charset='utf-8'>")
    html.append("<title>Sector ETF Heatmap</title>")
    # [CSS FIX]
    # 1. Removed 'table-layout: fixed' to allow auto-sizing based on content.
    # 2. Added specific style for the first column (Ticker/Name) to ensure it's not wrapped or squashed.
    html.append("""<style>
        body { margin: 0; padding: 20px; background: #0b0e14; font-family: Arial, sans-serif; color: #e2e8f0; }
        table { border-collapse: collapse; font-size: 13px; width: 100%; margin-top: 10px; }
        th, td { border: 1px solid #333; padding: 8px; text-align: center; white-space: nowrap; }
        th { background: #1e293b; color: #f8fafc; position: sticky; top: 0; z-index: 10; font-weight: bold; cursor: pointer; }
        td { background: #131722; color: #e2e8f0; }

        /* Fix for first column (Ticker Name) size */
        td:first-child, th:first-child {
            text-align: left;
            min-width: 150px;
            width: auto;
            white-space: nowrap;
        }

        h2 { text-align: center; margin-bottom: 5px; color: #f8fafc; }
        .timestamp { text-align: center; color: #94a3b8; font-size: 0.9em; margin-bottom: 20px; }
    </style></head><body>""")

    html.append("<h2>Sector ETF Heatmap</h2>")
    html.append(f"<div class='timestamp'>Last Updated: {timestamp_str} (UTC)</div>")

    html.append("<table>")
    html.append("<thead><tr><th>Ticker</th><th>Industry</th><th>ZS Live</th><th>ZS 5D</th>")
    for col in df_pct.columns: html.append(f"<th>{col}</th>")
    html.append("</tr></thead><tbody>")

    # Iterate through the DataFrame (which is passed in ALREADY SORTED)
    for idx, row in df_pct.iterrows():
        tk = str(idx)
        # Use simple white for text in first two columns
        html.append(f"<tr><td>{ticker_to_colb.get(tk, tk)}</td><td>{ticker_to_cola.get(tk, '')}</td>")

        # ZS Live
        vl = zs_live.loc[tk] if (zs_live is not None and tk in zs_live.index) else None
        bg_l = f"background-color:{color_for_zs_live(float(vl))}; color: white;" if (
                vl is not None and not pd.isna(vl) and color_for_zs_live(float(vl))) else ""
        html.append(
            f"<td style='{bg_l}'>{float(vl):+.2f}</td>" if (vl is not None and not pd.isna(vl)) else "<td></td>")

        # ZS 5D
        v5 = zs_5d.loc[tk] if (zs_5d is not None and tk in zs_5d.index) else None
        bg_5 = f"background-color:{color_for_zs_live(float(v5))}; color: white;" if (
                v5 is not None and not pd.isna(v5) and color_for_zs_live(float(v5))) else ""
        html.append(
            f"<td style='{bg_5}'>{float(v5):+.2f}</td>" if (v5 is not None and not pd.isna(v5)) else "<td></td>")

        for val in row.values:
            if pd.isna(val):
                html.append("<td></td>")
            else:
                clr = color_from_cmap_percent(float(val), cmap)
                # Ensure text is readable on colored background (black usually safe on heatmap colors)
                html.append(f"<td style='background-color:{clr}; color: black;'>{val:+.2f}%</td>")
        html.append("</tr>")
    html.append("</tbody></table></body></html>")

    with open(html_path, "w", encoding="utf-8") as f:
        f.write("\n".join(html))


def color_from_cmap_percent(val_pct: Optional[float], cmap: LinearSegmentedColormap) -> Optional[str]:
    if val_pct is None or np.isnan(val_pct): return None
    v = max(-10.0, min(10.0, float(val_pct)))
    return to_hex(cmap((v + 10.0) / 20.0), keep_alpha=False)


def _hex_to_rgb(hex_color: str) -> tuple:
    hex_color = hex_color.lstrip("#")
    return tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))


def _rgb_to_hex(rgb: tuple) -> str: return "#{:02X}{:02X}{:02X}".format(*rgb)


def color_for_zs_live(value: Optional[float], max_abs: float = 1.0) -> Optional[str]:
    if value is None or np.isnan(value) or max_abs <= 0: return None
    val = float(value)
    intensity = min(abs(val) / max_abs, 1.0)
    if intensity == 0: return None
    base_color = "#0FA84C" if val > 0 else "#C62828"
    rgb_base = _hex_to_rgb(base_color)
    rgb_white = (255, 255, 255)
    blended = tuple(int(round(rgb_white[i] + (rgb_base[i] - rgb_white[i]) * intensity)) for i in range(3))
    return _rgb_to_hex(blended)


def to_excel_with_fills(df_pct, xlsx_path, sheet_name="Heatmap", ticker_to_colb=None, ticker_to_cola=None, zs_live=None,
                        zs_5d=None):
    if ticker_to_colb is None: ticker_to_colb = {}
    if ticker_to_cola is None: ticker_to_cola = {}
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    df_for_excel = (df_pct.copy() / 100.0)
    df_for_excel.insert(0, "Ticker", [ticker_to_colb.get(str(i), str(i)) for i in df_for_excel.index])
    df_for_excel.insert(1, "Industry", [ticker_to_cola.get(str(i), "") for i in df_for_excel.index])

    zsl = zs_live.reindex(df_pct.index) if zs_live is not None else pd.Series(np.nan, index=df_pct.index)
    zs5 = zs_5d.reindex(df_pct.index) if zs_5d is not None else pd.Series(np.nan, index=df_pct.index)
    df_for_excel.insert(2, "ZS Live", pd.to_numeric(zsl, errors='coerce'))
    df_for_excel.insert(3, "ZS 5D", pd.to_numeric(zs5, errors='coerce'))

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_for_excel.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    max_row, max_col = ws.max_row, ws.max_column

    cmap = build_colormap()

    for r in range(2, max_row + 1):
        for c, val in [(3, ws.cell(r, 3).value), (4, ws.cell(r, 4).value)]:
            if val is not None and isinstance(val, (int, float)):
                ws.cell(r, c).number_format = "0.00;-0.00"
                clr = color_for_zs_live(val)
                if clr: ws.cell(r, c).fill = PatternFill(start_color=clr.replace("#", ""),
                                                         end_color=clr.replace("#", ""), fill_type="solid")

        for c in range(5, max_col + 1):
            cell = ws.cell(r, c)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
                pct = cell.value * 100.0
                hex_c = color_from_cmap_percent(pct, cmap)
                if hex_c: cell.fill = PatternFill(start_color=hex_c.replace("#", ""), end_color=hex_c.replace("#", ""),
                                                  fill_type="solid")

    wb.save(xlsx_path)


def generate_etf_sector_heatmap(output_html_path: str, output_excel_path: Optional[str] = None) -> str:
    output_html = Path(output_html_path).resolve()
    output_html.parent.mkdir(parents=True, exist_ok=True)
    output_excel = Path(output_excel_path).resolve() if output_excel_path else output_html.with_suffix(".xlsx")

    excel_path = BASE_DIR / "sector_etf_stock_list.xlsx"
    print(f"[INFO] Using Input Excel: {excel_path}")

    print("--- Calculating Volatility ---")
    update_volatility_360d_column(str(excel_path))

    print("--- Generating Heatmap ---")
    meta_df = pd.read_excel(excel_path)
    ticker_col = detect_ticker_column(meta_df)
    tickers = read_tickers_from_excel(str(excel_path), df=meta_df)
    print(f"Tickers found: {len(tickers)}")

    prices = download_prices(tickers, period="3mo")
    returns_pct = compute_daily_returns(prices)
    df_view = latest_20_days_transposed(returns_pct)

    ticker_to_colb = read_column_b_mapping(str(excel_path), df=meta_df)
    ticker_to_cola = read_column_a_mapping(str(excel_path), df=meta_df)
    vol_series = prepare_vol_series(meta_df, ticker_col, df_view.index)
    mu_series = prepare_mu_series(meta_df, ticker_col, df_view.index)
    zs_live = compute_zs_live(df_view, vol_series)
    zs_5d = compute_zs_5d(df_view, vol_series, mu_series)

    # --- [NEW] SORTING BY ZS LIVE ---
    # Create a temporary dataframe column for sorting
    # Note: zs_live has the same index as df_view (tickers)
    df_view["_temp_sort_zs"] = zs_live
    # Sort descending (highest ZS first), put NaNs at the bottom
    df_view = df_view.sort_values(by="_temp_sort_zs", ascending=False, na_position='last')
    # Remove temporary column
    df_view = df_view.drop(columns=["_temp_sort_zs"])
    # --------------------------------

    to_html_heatmap(df_view, str(output_html), ticker_to_colb, ticker_to_cola, zs_live, zs_5d)
    # Excel can remain static or dynamic, usually static is fine, but we'll focus on HTML for the web app
    to_excel_with_fills(df_view, str(output_excel), "Heatmap", ticker_to_colb, ticker_to_cola, zs_live, zs_5d)

    return str(output_html)


def main():
    # [KEY CHANGE] Generate dynamic filename with timestamp in the current directory
    file_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"sector_etf_heatmap_{file_timestamp}.html"

    # Save to the script's directory (BASE_DIR)
    output_html = BASE_DIR / output_filename
    output_xlsx = BASE_DIR / "sector_etf_heatmap.xlsx"  # Excel can stay static if not served via web

    print(f"[INFO] Generating report: {output_html}")
    generate_etf_sector_heatmap(str(output_html), str(output_xlsx))
    print(f"[SUCCESS] Heatmap generated: {output_filename}")


if __name__ == "__main__":
    main()